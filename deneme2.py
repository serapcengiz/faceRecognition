import tkinter as tk
from datetime import date
from openpyxl import Workbook, load_workbook
from PIL import ImageTk, Image
from tkinter import filedialog
import shutil
import os
import cv2
import face_recognition

from openpyxl import load_workbook

def kamera_ac():
    # Images klasöründeki fotoğrafları yüz tanıma için yükle
    yuzler = []
    etiketler = []
    dosya_yolu = 'images'

    for dosya_adi in os.listdir(dosya_yolu):
        if dosya_adi.endswith(".jpg") or dosya_adi.endswith(".png"):
            # Fotoğrafın tam dosya yolu
            foto_yolu = os.path.join(dosya_yolu, dosya_adi)

            # Yüzü bul
            yuz = face_recognition.load_image_file(foto_yolu)
            yuzler.append(face_recognition.face_encodings(yuz)[0])

            # Etiketi al (dosya adı)
            etiket = os.path.splitext(dosya_adi)[0]
            etiketler.append(etiket)

    # Kamera yakalama nesnesini oluştur
    kamera = cv2.VideoCapture(0)

    eslesen_foto_yolu = None  # Eşleşen fotoğraf yolunu saklamak için değişken

    while True:
        # Kameradan bir kare al
        ret, kare = kamera.read()

        # Kareyi küçült (hız için)
        kare = cv2.resize(kare, (0, 0), fx=1, fy=1)

        # Kareyi RGB formata dönüştür
        rgb_kare = kare[:, :, ::-1]

        # Yüzleri bul
        tespit_edilen_yuzler = face_recognition.face_locations(rgb_kare)
        tespit_edilen_yuz_tanima_kodlari = face_recognition.face_encodings(rgb_kare, tespit_edilen_yuzler)

        for tespit_edilen_yuz_tanima_kodu, (x, y, w, h) in zip(tespit_edilen_yuz_tanima_kodlari, tespit_edilen_yuzler):
            # Yüzü tanı
            eslesme = face_recognition.compare_faces(yuzler, tespit_edilen_yuz_tanima_kodu)
            isim = "Bilinmiyor"  # Tanınmayan yüzler için varsayılan ad

            if True in eslesme:
                indeks = eslesme.index(True)
                isim = etiketler[indeks]

                # Eşleşen fotoğraf yolunu al
                eslesen_foto_yolu = os.path.join(dosya_yolu, etiketler[indeks] + ".jpg")

                # Kareye yüzün çevresini çiz
                cv2.rectangle(kare, (h, x), (w, y), (0, 255, 0), 2)

                # Yüzün adını yaz
                cv2.putText(kare, isim, (h, x - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

        # Kareyi göster
        cv2.imshow('Kamera', kare)

        # Çıkış için q tuşuna basılmasını bekle
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Kamera yakalamasını serbest bırak
    kamera.release()
    cv2.destroyAllWindows()

    if eslesen_foto_yolu:
        print("Eşleşen fotoğraf yol:", eslesen_foto_yolu)
    else:
        print("Eşleşen fotoğraf bulunamadı.")

    dosya_adi = "ogrenciler.xlsx"
    workbook = load_workbook(dosya_adi)
    sheet = workbook.active

    eslesen_row = None
    eslesen_index = None

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[4] == eslesen_foto_yolu:
            print("Eşleşen satır:")
            print(row)
            row_list = list(row)
            if row_list[3] is None:
                row_list[3] = "geldi"
            sheet.append(row_list)
            eslesen_row = row
            eslesen_index = i
            break

    if eslesen_row:
        sheet.delete_rows(eslesen_index)

    workbook.save(dosya_adi)




def get_gunun_tarihi():
    today = date.today()
    gunun_tarihi = today.strftime("%d.%m.%Y")
    return gunun_tarihi


def fotoyu_sec():
    global dosya_yolu
    dosya_yolu = filedialog.askopenfilename(initialdir="/", title="Fotoğraf Seç",
                                            filetypes=(("JPEG dosyaları", "*.jpg"), ("PNG dosyaları", "*.png")))
    if dosya_yolu:
        img = Image.open(dosya_yolu)
        img = img.resize((100, 100))
        photo = ImageTk.PhotoImage(img)
        foto_label.configure(image=photo)
        foto_label.image = photo
        foto_label.pack()


def kaydet():
    ad = ad_entry.get()
    soyad = soyad_entry.get()
    numara = numara_entry.get()

    dosya_adi = "ogrenciler.xlsx"
    klasor_adi = "images"

    try:
        workbook = load_workbook(dosya_adi)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Ad"
        sheet["B1"] = "Soyad"
        sheet["C1"] = "Numara"
        sheet["D1"] = get_gunun_tarihi()
        sheet["E1"] = "Fotoğraf"


    # Mevcut verileri kontrol et
    mevcut_veriler = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        mevcut_veriler.append((row[0], row[1], row[2]))

    # Yeni öğrenci bilgileri mevcut verilerde yoksa ekle
    if (ad, soyad, numara) not in mevcut_veriler:
        satir_sayisi = sheet.max_row + 1
        sheet.cell(row=satir_sayisi, column=1, value=ad)
        sheet.cell(row=satir_sayisi, column=2, value=soyad)
        sheet.cell(row=satir_sayisi, column=3, value=numara)

        if dosya_yolu:
            # Yeni bir isim oluştur
            yeni_isim = ad + "_" + soyad + os.path.splitext(dosya_yolu)[1]
            hedef_yol = os.path.join(klasor_adi, yeni_isim)

            # Dosyayı images klasörüne kopyala
            shutil.copy(dosya_yolu, hedef_yol)

            # Dosya yolu bilgisini Excel'e kaydet
            sheet.cell(row=satir_sayisi, column=5, value=hedef_yol)

        # Günün tarihini al
        gunun_tarihi = date.today().strftime("%d.%m.%Y")
        sheet.cell(row=1, column=4, value=gunun_tarihi)

        print("Öğrenci bilgileri başarıyla kaydedildi.")

        # Alanları sıfırla
        ad_entry.delete(0, 'end')
        soyad_entry.delete(0, 'end')
        numara_entry.delete(0, 'end')
        foto_label.configure(image=None)

    else:
        print("Bu öğrenci bilgileri zaten mevcut.")

    # Dosyayı kaydet
    workbook.save(dosya_adi)


root = tk.Tk()
root.title("Öğrenci Kayıt Sistemi")

ad_label = tk.Label(root, text="Ad:")
ad_label.pack()

ad_entry = tk.Entry(root)
ad_entry.pack()

soyad_label = tk.Label(root, text="Soyad:")
soyad_label.pack()
soyad_entry = tk.Entry(root)
soyad_entry.pack()

numara_label = tk.Label(root, text="Numara:")
numara_label.pack()
numara_entry = tk.Entry(root)
numara_entry.pack()

foto_button = tk.Button(root, text="Fotoğraf Seç", command=fotoyu_sec)
foto_button.pack()

foto_label = tk.Label(root)
foto_label.pack()

kaydet_button = tk.Button(root, text="Kaydet", command=kaydet)
kaydet_button.pack()

kamera_button = tk.Button(root, text="Kamerayı Aç", command=kamera_ac)
kamera_button.pack()

kamera_label = tk.Label(root)
kamera_label.pack()

# İmages klasörünü oluştur
if not os.path.exists("venv/images"):
    os.makedirs("venv/images")

root.mainloop()
